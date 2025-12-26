from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
import shutil

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

SECRET_KEY = os.environ.get('JWT_SECRET', 'sanwariya-tms-secret-key-2026')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7

POD_UPLOAD_DIR = ROOT_DIR / "pod_uploads"
POD_UPLOAD_DIR.mkdir(exist_ok=True)

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    email: EmailStr
    name: str
    role: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class TripCreate(BaseModel):
    loading_date: str
    unloading_date: Optional[str] = None
    vehicle_number: str
    driver_mobile: str
    is_own_vehicle: bool
    motor_owner_name: Optional[str] = None
    motor_owner_mobile: Optional[str] = None
    gadi_bhada: Optional[float] = None
    gadi_advance: Optional[float] = None
    party_name: str
    party_mobile: str
    party_freight: float
    party_advance: Optional[float] = None
    tds: Optional[float] = None
    from_location: str
    to_location: str
    weight: Optional[str] = None
    himmali: Optional[str] = None
    remarks: Optional[str] = None
    status: str = "Loaded"

class Trip(BaseModel):
    model_config = ConfigDict(extra="ignore")
    trip_id: str
    loading_date: str
    unloading_date: Optional[str] = None
    vehicle_number: str
    driver_mobile: str
    is_own_vehicle: bool
    motor_owner_name: Optional[str] = None
    motor_owner_mobile: Optional[str] = None
    gadi_bhada: Optional[float] = None
    gadi_advance: Optional[float] = None
    gadi_balance: Optional[float] = None
    party_name: str
    party_mobile: str
    party_freight: float
    party_advance: Optional[float] = None
    party_balance: Optional[float] = None
    tds: Optional[float] = None
    from_location: str
    to_location: str
    weight: Optional[str] = None
    himmali: Optional[str] = None
    remarks: Optional[str] = None
    status: str
    settlement_status: str = "Pending"
    pod_filename: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TripUpdate(BaseModel):
    loading_date: Optional[str] = None
    unloading_date: Optional[str] = None
    vehicle_number: Optional[str] = None
    driver_mobile: Optional[str] = None
    is_own_vehicle: Optional[bool] = None
    motor_owner_name: Optional[str] = None
    motor_owner_mobile: Optional[str] = None
    gadi_bhada: Optional[float] = None
    gadi_advance: Optional[float] = None
    party_name: Optional[str] = None
    party_mobile: Optional[str] = None
    party_freight: Optional[float] = None
    party_advance: Optional[float] = None
    tds: Optional[float] = None
    from_location: Optional[str] = None
    to_location: Optional[str] = None
    weight: Optional[str] = None
    himmali: Optional[str] = None
    remarks: Optional[str] = None
    status: Optional[str] = None
    settlement_status: Optional[str] = None

class PartyAnalytics(BaseModel):
    party_name: str
    party_mobile: str
    total_trips: int
    total_freight: float
    total_paid: float
    outstanding_balance: float

class MotorOwnerAnalytics(BaseModel):
    motor_owner_name: str
    motor_owner_mobile: str
    total_trips: int
    total_bhada: float
    total_paid: float
    outstanding_balance: float

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await db.users.find_one({"email": email}, {"_id": 0, "password": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return User(**user)
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_admin_user(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

async def generate_trip_id() -> str:
    current_year = datetime.now(timezone.utc).year
    last_trip = await db.trips.find_one(
        {"trip_id": {"$regex": f"^{current_year}_"}},
        sort=[("created_at", -1)]
    )
    if last_trip:
        last_number = int(last_trip["trip_id"].split("_")[1])
        new_number = last_number + 1
    else:
        new_number = 1
    return f"{current_year}_{new_number}"

@api_router.post("/auth/register", response_model=User)
async def register_user(user_data: UserCreate):
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed_password = hash_password(user_data.password)
    user_dict = user_data.model_dump()
    user_dict["password"] = hashed_password
    user_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.users.insert_one(user_dict)
    return User(
        email=user_data.email,
        name=user_data.name,
        role=user_data.role,
        created_at=datetime.now(timezone.utc)
    )

@api_router.post("/auth/login", response_model=Token)
async def login(user_data: UserLogin):
    user = await db.users.find_one({"email": user_data.email})
    if not user or not verify_password(user_data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    access_token = create_access_token(data={"sub": user["email"]})
    user_obj = User(
        email=user["email"],
        name=user["name"],
        role=user["role"],
        created_at=datetime.fromisoformat(user["created_at"]) if isinstance(user["created_at"], str) else user["created_at"]
    )
    return Token(access_token=access_token, token_type="bearer", user=user_obj)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

@api_router.post("/trips", response_model=Trip)
async def create_trip(trip_data: TripCreate, current_user: User = Depends(get_current_user)):
    trip_id = await generate_trip_id()
    
    gadi_balance = None
    if not trip_data.is_own_vehicle and trip_data.gadi_bhada is not None:
        gadi_balance = trip_data.gadi_bhada - (trip_data.gadi_advance or 0)
    
    party_balance = trip_data.party_freight - (trip_data.party_advance or 0)
    
    trip_dict = trip_data.model_dump()
    trip_dict["trip_id"] = trip_id
    trip_dict["gadi_balance"] = gadi_balance
    trip_dict["party_balance"] = party_balance
    trip_dict["created_by"] = current_user.email
    trip_dict["created_at"] = datetime.now(timezone.utc)
    
    await db.trips.insert_one(trip_dict)
    
    # Convert created_at back to datetime object for the response
    trip_dict["created_at"] = datetime.now(timezone.utc)
    return Trip(**trip_dict)

@api_router.get("/trips", response_model=List[Trip])
async def get_trips(current_user: User = Depends(get_current_user)):
    query = {}
    projection = {"_id": 0}
    
    if current_user.role == "motor_owner":
        query["motor_owner_mobile"] = current_user.email
        projection["party_freight"] = 0
    
    trips = await db.trips.find(query, projection).sort("created_at", -1).to_list(1000)
    
    for trip in trips:
        if isinstance(trip.get("created_at"), str):
            trip["created_at"] = datetime.fromisoformat(trip["created_at"])
    
    return trips

@api_router.get("/trips/{trip_id}", response_model=Trip)
async def get_trip(trip_id: str, current_user: User = Depends(get_current_user)):
    trip = await db.trips.find_one({"trip_id": trip_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if current_user.role == "motor_owner" and trip.get("motor_owner_mobile") != current_user.email:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if isinstance(trip.get("created_at"), str):
        trip["created_at"] = datetime.fromisoformat(trip["created_at"])
    
    return Trip(**trip)

@api_router.put("/trips/{trip_id}", response_model=Trip)
async def update_trip(trip_id: str, trip_data: TripUpdate, current_user: User = Depends(get_current_user)):
    existing_trip = await db.trips.find_one({"trip_id": trip_id})
    if not existing_trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if current_user.role == "user" and existing_trip.get("status") == "Completed":
        raise HTTPException(status_code=403, detail="Cannot edit completed trips")
    
    update_dict = {k: v for k, v in trip_data.model_dump(exclude_unset=True).items() if v is not None}
    
    if "gadi_bhada" in update_dict or "gadi_advance" in update_dict:
        gadi_bhada = update_dict.get("gadi_bhada", existing_trip.get("gadi_bhada"))
        gadi_advance = update_dict.get("gadi_advance", existing_trip.get("gadi_advance", 0))
        if gadi_bhada is not None:
            update_dict["gadi_balance"] = gadi_bhada - gadi_advance
    
    if "party_freight" in update_dict or "party_advance" in update_dict:
        party_freight = update_dict.get("party_freight", existing_trip.get("party_freight"))
        party_advance = update_dict.get("party_advance", existing_trip.get("party_advance", 0))
        update_dict["party_balance"] = party_freight - party_advance
    
    await db.trips.update_one({"trip_id": trip_id}, {"$set": update_dict})
    
    updated_trip = await db.trips.find_one({"trip_id": trip_id}, {"_id": 0})
    if isinstance(updated_trip.get("created_at"), str):
        updated_trip["created_at"] = datetime.fromisoformat(updated_trip["created_at"])
    
    return Trip(**updated_trip)

@api_router.delete("/trips/{trip_id}")
async def delete_trip(trip_id: str, current_user: User = Depends(get_admin_user)):
    result = await db.trips.delete_one({"trip_id": trip_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Trip not found")
    return {"message": "Trip deleted successfully"}

@api_router.post("/trips/{trip_id}/pod")
async def upload_pod(trip_id: str, file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    trip = await db.trips.find_one({"trip_id": trip_id})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    file_extension = file.filename.split(".")[-1]
    filename = f"{trip_id}_pod.{file_extension}"
    file_path = POD_UPLOAD_DIR / filename
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    await db.trips.update_one({"trip_id": trip_id}, {"$set": {"pod_filename": filename}})
    
    return {"message": "POD uploaded successfully", "filename": filename}

@api_router.get("/trips/{trip_id}/pod")
async def download_pod(trip_id: str, current_user: User = Depends(get_current_user)):
    trip = await db.trips.find_one({"trip_id": trip_id}, {"_id": 0})
    if not trip or not trip.get("pod_filename"):
        raise HTTPException(status_code=404, detail="POD not found")
    
    file_path = POD_UPLOAD_DIR / trip["pod_filename"]
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="POD file not found")
    
    return FileResponse(file_path, filename=trip["pod_filename"])

@api_router.get("/analytics/parties", response_model=List[PartyAnalytics])
async def get_party_analytics(current_user: User = Depends(get_admin_user)):
    pipeline = [
        {"$group": {
            "_id": {"name": "$party_name", "mobile": "$party_mobile"},
            "total_trips": {"$sum": 1},
            "total_freight": {"$sum": "$party_freight"},
            "total_paid": {"$sum": "$party_advance"},
        }},
        {"$project": {
            "_id": 0,
            "party_name": "$_id.name",
            "party_mobile": "$_id.mobile",
            "total_trips": 1,
            "total_freight": 1,
            "total_paid": 1,
            "outstanding_balance": {"$subtract": ["$total_freight", "$total_paid"]}
        }}
    ]
    
    results = await db.trips.aggregate(pipeline).to_list(1000)
    return results

@api_router.get("/analytics/motor-owners", response_model=List[MotorOwnerAnalytics])
async def get_motor_owner_analytics(current_user: User = Depends(get_admin_user)):
    pipeline = [
        {"$match": {"is_own_vehicle": False, "motor_owner_name": {"$ne": None}}},
        {"$group": {
            "_id": {"name": "$motor_owner_name", "mobile": "$motor_owner_mobile"},
            "total_trips": {"$sum": 1},
            "total_bhada": {"$sum": "$gadi_bhada"},
            "total_paid": {"$sum": "$gadi_advance"},
        }},
        {"$project": {
            "_id": 0,
            "motor_owner_name": "$_id.name",
            "motor_owner_mobile": "$_id.mobile",
            "total_trips": 1,
            "total_bhada": 1,
            "total_paid": 1,
            "outstanding_balance": {"$subtract": ["$total_bhada", "$total_paid"]}
        }}
    ]
    
    results = await db.trips.aggregate(pipeline).to_list(1000)
    return results

@api_router.get("/export/trips")
async def export_trips(month: Optional[int] = None, year: Optional[int] = None, current_user: User = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    query = {}
    if month and year:
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
        query["loading_date"] = {"$gte": start_date, "$lt": end_date}
    
    trips = await db.trips.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Trips"
    
    headers = [
        "Trip ID", "Loading Date", "Unloading Date", "Vehicle No", "Driver No",
        "Motor Owner Name", "Motor Owner Mobile", "Gadi Bhada", "Gadi Advance", "Gadi Balance",
        "Party Name", "Party Mobile", "Party Freight", "Party Advance", "Party Balance",
        "TDS", "From", "To", "Weight", "Himmali", "Remarks", "Status"
    ]
    
    header_fill = PatternFill(start_color="E89448", end_color="E89448", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row, trip in enumerate(trips, 2):
        ws.cell(row=row, column=1, value=trip.get("trip_id"))
        ws.cell(row=row, column=2, value=trip.get("loading_date"))
        ws.cell(row=row, column=3, value=trip.get("unloading_date"))
        ws.cell(row=row, column=4, value=trip.get("vehicle_number"))
        ws.cell(row=row, column=5, value=trip.get("driver_mobile"))
        ws.cell(row=row, column=6, value=trip.get("motor_owner_name"))
        ws.cell(row=row, column=7, value=trip.get("motor_owner_mobile"))
        ws.cell(row=row, column=8, value=trip.get("gadi_bhada"))
        ws.cell(row=row, column=9, value=trip.get("gadi_advance"))
        ws.cell(row=row, column=10, value=trip.get("gadi_balance"))
        ws.cell(row=row, column=11, value=trip.get("party_name"))
        ws.cell(row=row, column=12, value=trip.get("party_mobile"))
        ws.cell(row=row, column=13, value=trip.get("party_freight"))
        ws.cell(row=row, column=14, value=trip.get("party_advance"))
        ws.cell(row=row, column=15, value=trip.get("party_balance"))
        ws.cell(row=row, column=16, value=trip.get("tds"))
        ws.cell(row=row, column=17, value=trip.get("from_location"))
        ws.cell(row=row, column=18, value=trip.get("to_location"))
        ws.cell(row=row, column=19, value=trip.get("weight"))
        ws.cell(row=row, column=20, value=trip.get("himmali"))
        ws.cell(row=row, column=21, value=trip.get("remarks"))
        ws.cell(row=row, column=22, value=trip.get("status"))
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    filename = f"trips_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = POD_UPLOAD_DIR / filename
    wb.save(filepath)
    
    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

@app.on_event("startup")
async def create_admin_user():
    admin_email = "shrisanwariyaroadlines@gmail.com"
    existing_admin = await db.users.find_one({"email": admin_email})
    if not existing_admin:
        admin_user = {
            "email": admin_email,
            "password": hash_password("Sanwariya_1228"),
            "name": "Admin",
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(admin_user)
        logger.info(f"Admin user created: {admin_email}")
